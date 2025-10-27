import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import copy
import glob
import os
import threading
import sys

# 預設值（今天年月）
default_tag = datetime.today().strftime("%Y%m")
user_input = {"value": None}

def ask_input():
    try:
        user_input["value"] = input("請輸入一或多個年月（例如 202405 202406），10 秒內未輸入則自動使用當月：").strip()
    except EOFError:
        user_input["value"] = ""

# 啟動輸入監聽執行緒
t = threading.Thread(target=ask_input)
t.daemon = True
t.start()
t.join(timeout=10)  # 最多等 10 秒

# 判斷結果
if user_input["value"]:
    raw_tags = user_input["value"].replace(",", " ").split()
    month_tags = [tag for tag in raw_tags if len(tag) == 6 and tag.isdigit()]
else:
    print(f"⏰ 超過 10 秒未輸入，自動使用 {default_tag}")
    month_tags = [default_tag]


# 資料來源檔案
data_file = "data.xlsx"
data_wb = load_workbook(data_file)
data_ws = data_wb["IM"]

# 取得所有現有案號（C欄，第3欄）
existing_case_ids = set()
for row in data_ws.iter_rows(min_row=2, min_col=3, max_col=3):
    val = row[0].value
    if val:
        existing_case_ids.add(str(val).strip())

# 參考格式用於複製樣式與公式
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}
al_formula_cell = data_ws[f"AL{ref_row}"]
am_formula_cell = data_ws[f"AM{ref_row}"]

total_new_rows = 0

# 處理多個報表檔案
for tag in month_tags:
    report_file = f"IM/{tag}_HL_Maintain_Report.xlsx"
    if not os.path.exists(report_file):
        print(f"❌ 找不到：{report_file}")
        continue

    print(f"🔄 處理報表：{report_file}")
    report_wb = load_workbook(report_file, data_only=True)
    report_ws = report_wb.active

    start_row = 2
    append_rows = []
    for row in report_ws.iter_rows(min_row=start_row):
        case_cell = row[2]  # C欄
        case_id_raw = str(case_cell.value).strip() if case_cell.value else ""
        if not case_id_raw or not case_id_raw.isdigit():
            continue

        if case_id_raw not in existing_case_ids:
            values = [cell.value for cell in row]
            if all(v is None for v in values):
                break
            append_rows.append(values)
            existing_case_ids.add(case_id_raw)  # 確保不重複新增

    print(f"   ➕ 發現 {len(append_rows)} 列新資料")
    total_new_rows += len(append_rows)

    # 寫入資料
    for row_data in append_rows:
        data_ws.append(row_data)
        new_row = data_ws.max_row
        data_ws.row_dimensions[new_row].height = ref_row_height

        for col_idx, value in enumerate(row_data, start=1):
            cell = data_ws.cell(row=new_row, column=col_idx)
            ref_cell = ref_cells.get(col_idx)
            if ref_cell:
                cell.font = copy.copy(ref_cell.font)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)

            if col_idx == 24 and value:
                try:
                    if isinstance(value, str):
                        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None
                    if dt:
                        cell.value = dt
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except Exception as e:
                    print(f"❗ 日期格式錯誤（欄{col_idx}）：{e}")

        # 複製公式
        for col_letter, formula_cell in [("AL", al_formula_cell), ("AM", am_formula_cell)]:
            target_cell = data_ws[f"{col_letter}{new_row}"]
            formula = formula_cell.value.replace(str(ref_row), str(new_row)) if formula_cell.data_type == "f" else formula_cell.value
            target_cell.value = formula
            target_cell.font = copy.copy(formula_cell.font)
            target_cell.alignment = copy.copy(formula_cell.alignment)
            target_cell.border = copy.copy(formula_cell.border)
            target_cell.number_format = formula_cell.number_format

# 儲存更新
data_wb.save(data_file)
print(f"✅ 更新完成，共加入 {total_new_rows} 筆資料")
