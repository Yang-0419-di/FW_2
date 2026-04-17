import os
import time
import copy
import threading
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers

# ========= 等待檔案 =========
def wait_file_ready(path, timeout=10):
    last_size = -1
    for _ in range(timeout):
        if os.path.exists(path):
            size = os.path.getsize(path)
            if size == last_size:
                return True
            last_size = size
        time.sleep(1)
    return False

# ========= 輸入年月 =========
def get_month_tags():
    default_tag = datetime.today().strftime("%Y%m")
    user_input = {"value": None}

    def ask():
        try:
            user_input["value"] = input("請輸入年月（例如 202405 202406）：").strip()
        except:
            user_input["value"] = ""

    t = threading.Thread(target=ask)
    t.daemon = True
    t.start()
    t.join(timeout=10)

    if user_input["value"]:
        raw = user_input["value"].replace(",", " ").split()
        return [x for x in raw if len(x) == 6 and x.isdigit()]
    else:
        print(f"⏰ 自動使用 {default_tag}")
        return [default_tag]

# ========= 處理單一 sheet =========
def process_sheet(ws, config, base_dir, month_tags):

    case_col = config["case_col"]
    report_name = config["report_name"]
    filter_func = config.get("filter_func")
    copy_formula = config.get("copy_formula", False)

    existing = set()
    for row in ws.iter_rows(min_row=2, min_col=case_col, max_col=case_col):
        if row[0].value:
            existing.add(str(row[0].value).strip())

    ref_row = ws.max_row
    ref_cells = {c.column: c for c in ws[ref_row]}
    ref_height = 21.66

    if copy_formula:
        al_formula = ws[f"AL{ref_row}"]
        am_formula = ws[f"AM{ref_row}"]

    total = 0

    for tag in month_tags:
        report_file = os.path.join(base_dir, "IM", f"{tag}_{report_name}.xlsx")

        if not wait_file_ready(report_file):
            print(f"❌ 找不到：{report_file}")
            continue

        print(f"🔄 處理：{report_file}")
        r_wb = load_workbook(report_file, data_only=True)
        r_ws = r_wb.active

        new_rows = []

        for row in r_ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if all(v is None for v in values):
                continue

            case_id = str(row[case_col - 1].value).strip() if row[case_col - 1].value else ""
            if not case_id.isdigit():
                continue

            if filter_func and not filter_func(row):
                continue

            if case_id not in existing:
                new_rows.append(values)
                existing.add(case_id)

        print(f"   ➕ 新增 {len(new_rows)} 筆")
        total += len(new_rows)

        for row_data in new_rows:
            ws.append(row_data)
            new_row = ws.max_row
            ws.row_dimensions[new_row].height = ref_height

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=new_row, column=col_idx)
                ref = ref_cells.get(col_idx)

                if ref:
                    cell.font = copy.copy(ref.font)
                    cell.alignment = copy.copy(ref.alignment)
                    cell.border = copy.copy(ref.border)
                    cell.fill = copy.copy(ref.fill)

                if col_idx == 24 and val:
                    try:
                        if isinstance(val, str):
                            dt = datetime.strptime(val.strip(), "%Y-%m-%d %H:%M:%S")
                        elif isinstance(val, datetime):
                            dt = val
                        else:
                            dt = None

                        if dt:
                            cell.value = dt
                            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    except Exception as e:
                        print(f"❗ 日期錯誤：{e}")

            if copy_formula:
                for col, f_cell in [("AL", al_formula), ("AM", am_formula)]:
                    target = ws[f"{col}{new_row}"]
                    formula = f_cell.value.replace(str(ref_row), str(new_row)) if f_cell.data_type == "f" else f_cell.value
                    target.value = formula

        r_wb.close()

    return total

# ========= 主程式 =========
base_dir = os.path.dirname(os.path.abspath(__file__))
month_tags = get_month_tags()

try:
    file_jobs = {
        os.path.join(base_dir, "data.xlsx"): [
            {
                "sheet": "IM",
                "case_col": 3,
                "report_name": "HL_Maintain_Report",
                "copy_formula": True
            },
            {
                "sheet": "MFP",
                "case_col": 2,
                "report_name": "Service_Count_Report",
                "filter_func": lambda r: str(r[18].value).strip().upper() == "O"
                                          and not (r[11].value and "萊爾富" in str(r[11].value))
            }
        ],

        os.path.join(base_dir, "MFP", "MFP.xlsx"): [
            {
                "sheet": "IM",
                "case_col": 3,
                "report_name": "HL_Maintain_Report",
                "copy_formula": True
            },
            {
                "sheet": "MFP",
                "case_col": 2,
                "report_name": "Service_Count_Report",
                "filter_func": lambda r: str(r[18].value).strip().upper() == "O"
                                          and not (r[11].value and "萊爾富" in str(r[11].value))
            }
        ],

        os.path.join(base_dir, "MFP", "output.xlsx"): [
            {
                "sheet": "IM",
                "case_col": 3,
                "report_name": "HL_Maintain_Report",
                "copy_formula": True
            },
            {
                "sheet": "MFP",
                "case_col": 2,
                "report_name": "Service_Count_Report",
                "filter_func": lambda r: str(r[18].value).strip().upper() == "O"
                                          and not (r[11].value and "萊爾富" in str(r[11].value))
            }
        ]
    }

    for file_path, jobs in file_jobs.items():

        if not wait_file_ready(file_path):
            print(f"❌ 主檔錯誤：{file_path}")
            sys.exit(1)

        print(f"\n📂 開始處理：{file_path}")
        wb = load_workbook(file_path)

        total_all = 0

        for job in jobs:
            ws = wb[job["sheet"]]
            total = process_sheet(ws, job, base_dir, month_tags)
            total_all += total

        tmp_path = file_path + ".tmp"

        wb.save(tmp_path)
        wb.close()

        # 等一下確保寫完
        time.sleep(1)

        # 原子替換（避免壞檔）
        os.replace(tmp_path, file_path)

        print(f"✅ 完成 {file_path}，共新增 {total_all} 筆\n")

    sys.exit(0)

except Exception as e:
    print(f"❌ 發生錯誤：{e}")
    sys.exit(1)