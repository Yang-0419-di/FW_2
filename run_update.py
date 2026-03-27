import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from datetime import datetime
import copy
import os
import threading
import sys
import time

# ========= 函式：等待檔案就緒 =========
def wait_file_ready(path, timeout=10):
    """確認檔案存在且大小穩定"""
    last_size = -1
    for _ in range(timeout):
        if os.path.exists(path):
            size = os.path.getsize(path)
            if size == last_size:
                return True
            last_size = size
        time.sleep(1)
    return False

# ========= 預設值與使用者輸入 =========
default_tag = datetime.today().strftime("%Y%m")
user_input = {"value": None}

def ask_input():
    try:
        user_input["value"] = input(
            "請輸入一或多個年月（例如 202405 202406），10 秒內未輸入則自動使用當月："
        ).strip()
    except EOFError:
        user_input["value"] = ""

t = threading.Thread(target=ask_input)
t.daemon = True
t.start()
t.join(timeout=10)

if user_input["value"]:
    raw_tags = user_input["value"].replace(",", " ").split()
    month_tags = [tag for tag in raw_tags if len(tag) == 6 and tag.isdigit()]
else:
    print(f"⏰ 超過 10 秒未輸入，自動使用 {default_tag}")
    month_tags = [default_tag]

# ========= 路徑設定 =========
base_dir = os.path.dirname(os.path.abspath(__file__))
data_file = os.path.join(base_dir, "MFP", "output.xlsx")

if not wait_file_ready(data_file):
    print(f"❌ 找不到主檔或檔案不穩定：{data_file}")
    sys.exit(1)

# ========= 載入主檔 =========
data_wb = load_workbook(data_file)
data_ws = data_wb["IM"]  # 主檔分頁 IM

# 取得現有案號（C欄）
existing_case_ids = set()
for row in data_ws.iter_rows(min_row=2, min_col=3, max_col=3):
    val = row[0].value
    if val:
        existing_case_ids.add(str(val).strip())

# 參考格式列（最後一列）
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}
al_formula_cell = data_ws[f"AL{ref_row}"]
am_formula_cell = data_ws[f"AM{ref_row}"]

total_new_rows = 0

# ========= 處理每個月份報表 =========
for tag in month_tags:
    report_file = os.path.join(base_dir, "IM", f"{tag}_HL_Maintain_Report.xlsx")
    if not wait_file_ready(report_file):
        print(f"❌ 找不到報表或檔案不穩定：{report_file}")
        continue

    print(f"🔄 處理報表：{report_file}")
    report_wb = load_workbook(report_file, data_only=True)
    report_ws = report_wb.active

    append_rows = []
    for row in report_ws.iter_rows(min_row=2):
        case_cell = row[2]  # C欄案號
        case_id_raw = str(case_cell.value).strip() if case_cell.value else ""
        if not case_id_raw or not case_id_raw.isdigit():
            continue

        if case_id_raw not in existing_case_ids:
            values = [cell.value for cell in row]
            if all(v is None for v in values):
                continue
            append_rows.append(values)
            existing_case_ids.add(case_id_raw)

    print(f"   ➕ 發現 {len(append_rows)} 列新資料")
    total_new_rows += len(append_rows)

    # 寫入主檔
    for row_data in append_rows:
        data_ws.append(row_data)
        new_row = data_ws.max_row
        data_ws.row_dimensions[new_row].height = ref_row_height

        for col_idx, value in enumerate(row_data, start=1):
            cell = data_ws.cell(row=new_row, column=col_idx)
            ref_cell = ref_cells.get(col_idx)
            if ref_cell:
                cell.font = copy.copy(ref_cell.font)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)

            # 第24欄為日期時間
            if col_idx == 24 and value:
                try:
                    if isinstance(value, str):
                        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None
                    if dt:
                        cell.value = dt
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except Exception as e:
                    print(f"❗ 日期格式錯誤（欄{col_idx}）：{e}")

        # 複製公式（AL、AM欄）
        for col_letter, formula_cell in [("AL", al_formula_cell), ("AM", am_formula_cell)]:
            target_cell = data_ws[f"{col_letter}{new_row}"]
            formula = formula_cell.value.replace(str(ref_row), str(new_row)) if formula_cell.data_type == "f" else formula_cell.value
            target_cell.value = formula
            target_cell.font = copy.copy(formula_cell.font)
            target_cell.alignment = copy.copy(formula_cell.alignment)
            target_cell.border = copy.copy(formula_cell.border)
            target_cell.number_format = formula_cell.number_format

# ========= 儲存主檔 =========
data_wb.save(data_file)
print(f"✅ 更新完成，共加入 {total_new_rows} 筆資料")