import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from datetime import datetime
import copy
import os
import threading
import sys

# 預設值（今天年月）
default_tag = datetime.today().strftime("%Y%m")
user_input = {"value": None}

def ask_input():
    try:
        user_input["value"] = input("請輸入一或多個年月（例如 202405 202406），10 秒內未輸入則自動使用當月：").strip()
    except EOFError:
        user_input["value"] = ""

# 啟動輸入監聽執行緒
t = threading.Thread(target=ask_input)
t.daemon = True
t.start()
t.join(timeout=10)  # 最多等 10 秒

# 判斷結果
if user_input["value"]:
    raw_tags = user_input["value"].replace(",", " ").split()
    month_tags = [tag for tag in raw_tags if len(tag) == 6 and tag.isdigit()]
else:
    print(f"⏰ 超過 10 秒未輸入，自動使用 {default_tag}")
    month_tags = [default_tag]


# ========= 正確處理路徑 =========

base_dir = os.path.dirname(os.path.abspath(__file__))  # run_MFP2_update.py 位置

# MFP 主檔路徑
data_file = os.path.join(base_dir, "MFP", "output.xlsx")
if not os.path.exists(data_file):
    print(f"❌ 找不到主檔：{data_file}")
    sys.exit(1)

# IM 月報表路徑由 base_dir + IM 子資料夾建構
# report_file = os.path.join(base_dir, "IM", f"{tag}_Service_Count_Report.xlsx")

# =================================


# 載入 MFP 主檔
data_wb = load_workbook(data_file)
data_ws = data_wb["MFP"]  # 主檔分頁 IM

# 取得所有現有案號（B欄 = 第2欄）
existing_case_ids = set()
for row in data_ws.iter_rows(min_row=2, min_col=2, max_col=2):
    val = row[0].value
    if val:
        existing_case_ids.add(str(val).strip())

# 參考格式列（最後一列）作為樣板
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}

total_new_rows = 0

# 逐月處理資料
for tag in month_tags:

    report_file = os.path.join(base_dir, "IM", f"{tag}_Service_Count_Report.xlsx")

    if not os.path.exists(report_file):
        print(f"❌ 找不到：{report_file}")
        continue

    print(f"🔄 處理報表：{report_file}")
    report_wb = load_workbook(report_file, data_only=True)
    report_ws = report_wb.active

    start_row = 2
    append_rows = []

    # 掃描報表每列，確保讀到第28欄(AB欄)
    for row in report_ws.iter_rows(min_row=start_row, max_col=28):

        ab_value = row[27].value  # AB 欄 = 第28欄
        if ab_value != 1:
            continue

        l_value = row[11].value  # L欄
        if l_value and "萊爾富" in str(l_value):
            continue

        case_cell = row[1]  # 案號 B 欄
        case_id_raw = str(case_cell.value).strip() if case_cell.value else ""

        if not case_id_raw or not case_id_raw.isdigit():
            continue

        # 若案號不存在於主檔 → 新增列
        if case_id_raw not in existing_case_ids:
            values = [cell.value for cell in row]

            if all(v is None for v in values):
                break

            append_rows.append(values)
            existing_case_ids.add(case_id_raw)

    print(f"   ➕ 發現 {len(append_rows)} 列新資料")
    total_new_rows += len(append_rows)

    # 新增資料到主檔
    for row_data in append_rows:
        data_ws.append(row_data)
        new_row = data_ws.max_row
        data_ws.row_dimensions[new_row].height = ref_row_height

        for col_idx, value in enumerate(row_data, start=1):
            cell = data_ws.cell(row=new_row, column=col_idx)
            ref_cell = ref_cells.get(col_idx)

            if ref_cell:
                cell.font = copy.copy(ref_cell.font)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)

            # 第24欄為「日期時間」欄位
            if col_idx == 24 and value:
                try:
                    if isinstance(value, str):
                        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None

                    if dt:
                        cell.value = dt
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                except Exception as e:
                    print(f"❗ 日期格式錯誤（欄{col_idx}）：{e}")


# 儲存更新後的主檔
data_wb.save(data_file)
print(f"✅ 更新完成，共加入 {total_new_rows} 筆資料")
