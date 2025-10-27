import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from datetime import datetime
import copy
import os

# 詢問年月，可輸入多個用空格或逗號分隔
input_tags = input("請輸入一或多個年月（例如 202405 202406），直接 Enter 使用當月：").strip()
if input_tags:
    raw_tags = input_tags.replace(",", " ").split()
    month_tags = [tag for tag in raw_tags if len(tag) == 6 and tag.isdigit()]
else:
    month_tags = [datetime.today().strftime("%Y%m")]

# 資料來源檔案與工作表
data_file = "data.xlsx"
data_wb = load_workbook(data_file)
data_ws = data_wb["MFP"]  # ✅ 分頁改為 MFP

# 取得所有現有案號（B欄 = 第2欄 = row[1]）
existing_case_ids = set()
for row in data_ws.iter_rows(min_row=2, min_col=2, max_col=2):
    val = row[0].value
    if val:
        existing_case_ids.add(str(val).strip())

# 參考格式列（最後一列）作為樣板
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}

total_new_rows = 0

# 逐一處理輸入的年月
for tag in month_tags:
    report_file = f"IM/{tag}_Service_Count_Report.xlsx"  # ✅ 檔案名稱格式
    if not os.path.exists(report_file):
        print(f"❌ 找不到：{report_file}")
        continue

    print(f"🔄 處理報表：{report_file}")
    report_wb = load_workbook(report_file, data_only=True)
    report_ws = report_wb.active

    start_row = 2
    append_rows = []

    # 掃描報表每列，確保讀取到第28欄(AB欄)
    for row in report_ws.iter_rows(min_row=start_row, max_col=28):
        ab_value = row[27].value  # AB欄 (第28欄，索引27)
        if ab_value != 1:
            continue
        
        l_value = row[11].value  # L欄 (第12欄，索引11)
        if l_value and "萊爾富" in str(l_value):
            continue  # L欄包含萊爾富就跳過
        
        case_cell = row[1]  # 案號B欄 (索引1)
        case_id_raw = str(case_cell.value).strip() if case_cell.value else ""
        if not case_id_raw or not case_id_raw.isdigit():
            continue
        
        if case_id_raw not in existing_case_ids:
            values = [cell.value for cell in row]
            if all(v is None for v in values):
                break
            append_rows.append(values)
            existing_case_ids.add(case_id_raw)
            
    print(f"   ➕ 發現 {len(append_rows)} 列新資料")
    total_new_rows += len(append_rows)

    # 新增資料到主檔
    for row_data in append_rows:
        data_ws.append(row_data)
        new_row = data_ws.max_row
        data_ws.row_dimensions[new_row].height = ref_row_height

        for col_idx, value in enumerate(row_data, start=1):
            cell = data_ws.cell(row=new_row, column=col_idx)
            ref_cell = ref_cells.get(col_idx)
            if ref_cell:
                cell.font = copy.copy(ref_cell.font)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)

            # 若第24欄為時間，處理格式與樣式
            if col_idx == 24 and value:
                try:
                    if isinstance(value, str):
                        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None
                    if dt:
                        cell.value = dt
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except Exception as e:
                    print(f"❗ 日期格式錯誤（欄{col_idx}）：{e}")

# 儲存更新後的 Excel
data_wb.save(data_file)
print(f"✅ 更新完成，共加入 {total_new_rows} 筆資料")
