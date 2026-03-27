import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from datetime import datetime
import copy
import os
import threading
import sys
import time

# ========= 函式：等待檔案就緒 =========
def wait_file_ready(path, timeout=10):
    """確認檔案存在且大小穩定"""
    last_size = -1
    for _ in range(timeout):
        if os.path.exists(path):
            size = os.path.getsize(path)
            if size == last_size:
                return True
            last_size = size
        time.sleep(1)
    return False

# ========= 路徑設定 =========
base_dir = os.path.dirname(os.path.abspath(__file__))
data_file = os.path.join(base_dir, "data.xlsx")

# 等待主檔就緒
if not wait_file_ready(data_file):
    print(f"❌ 主檔不存在或無法穩定讀取：{data_file}")
    sys.exit(1)

# ========= 預設值（今天年月）與輸入 =========
default_tag = datetime.today().strftime("%Y%m")
user_input = {"value": None}

def ask_input():
    try:
        user_input["value"] = input(
            "請輸入一或多個年月（例如 202405 202406），10 秒內未輸入則自動使用當月："
        ).strip()
    except EOFError:
        user_input["value"] = ""

t = threading.Thread(target=ask_input)
t.daemon = True
t.start()
t.join(timeout=10)  # 最多等 10 秒

if user_input["value"]:
    raw_tags = user_input["value"].replace(",", " ").split()
    month_tags = [tag for tag in raw_tags if len(tag) == 6 and tag.isdigit()]
else:
    print(f"⏰ 超過 10 秒未輸入，自動使用 {default_tag}")
    month_tags = [default_tag]

# ========= 載入 Excel 主檔 =========
data_wb = load_workbook(data_file)
data_ws = data_wb["MFP"]  # 主檔分頁 MFP

# 取得所有現有案號（B欄）
existing_case_ids = set()
for row in data_ws.iter_rows(min_row=2, min_col=2, max_col=2):
    val = row[0].value
    if val:
        existing_case_ids.add(str(val).strip())

# 參考格式列（最後一列）作為樣板
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}

total_new_rows = 0

# ========= 逐月處理報表 =========
for tag in month_tags:
    report_file = os.path.join(base_dir, "IM", f"{tag}_Service_Count_Report.xlsx")

    if not wait_file_ready(report_file):
        print(f"❌ 報表不存在或無法穩定讀取：{report_file}")
        continue

    print(f"🔄 處理報表：{report_file}")
    report_wb = load_workbook(report_file, data_only=True)
    report_ws = report_wb.active

    start_row = 2
    append_rows = []

    for row in report_ws.iter_rows(min_row=start_row, max_col=28):
        s_value = row[18].value  # S欄
        if str(s_value).strip().upper() != "O":
            continue

        l_value = row[11].value  # L欄
        if l_value and "萊爾富" in str(l_value):
            continue

        case_cell = row[1]  # B欄案號
        case_id_raw = str(case_cell.value).strip() if case_cell.value else ""
        if not case_id_raw or not case_id_raw.isdigit():
            continue

        if case_id_raw not in existing_case_ids:
            values = [cell.value for cell in row]
            if all(v is None for v in values):
                continue  # 空列跳過
            append_rows.append(values)
            existing_case_ids.add(case_id_raw)

    print(f"   ➕ 發現 {len(append_rows)} 列新資料")
    total_new_rows += len(append_rows)

    # 新增資料到主檔
    for row_data in append_rows:
        data_ws.append(row_data)
        new_row = data_ws.max_row
        data_ws.row_dimensions[new_row].height = ref_row_height

        for col_idx, value in enumerate(row_data, start=1):
            cell = data_ws.cell(row=new_row, column=col_idx)
            ref_cell = ref_cells.get(col_idx)
            if ref_cell:
                cell.font = copy.copy(ref_cell.font)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)

            # 第24欄為「日期時間」欄位
            if col_idx == 24 and value:
                try:
                    if isinstance(value, str):
                        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None
                    if dt:
                        cell.value = dt
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except Exception as e:
                    print(f"❗ 日期格式錯誤（欄{col_idx}）：{e}")

# ========= 儲存更新後主檔 =========
data_wb.save(data_file)
print(f"✅ 更新完成，共加入 {total_new_rows} 筆資料")